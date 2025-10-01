# ctdmate/tools/reg_rag.py
from __future__ import annotations
from pathlib import Path
from typing import Dict, List, Any, Optional
import re
import openpyxl

# config
try:
    from ctdmate.app import config as CFG
except Exception:
    from ..app import config as CFG  # type: ignore

# RAG + 정규화 도구
try:
    from ctdmate.rag.mfds_rag import MFDSRAGTool
    from ctdmate.rag.glossary_rag import GlossaryRAGTool
    from ctdmate.rag.term_normalizer import TermNormalizer
except Exception:
    from ..rag.mfds_rag import MFDSRAGTool  # type: ignore
    from ..rag.glossary_rag import GlossaryRAGTool  # type: ignore
    from ..rag.term_normalizer import TermNormalizer  # type: ignore

SHEET_TO_MODULE = {
    "TM_5_M2_3_QOS": "M2.3",
    "TM_5_M2_4_Nonclinical_Ove": "M2.4",
    "TM_5_M2_5_Clinical_Overvi": "M2.5",
    "TM_5_M2_6_Nonclinical_Sum": "M2.6",
    "TM_5_M2_7_Clinical_Summar": "M2.7",
    "TM_5_Admin_Labeling_KR": "M1",
    "TM_5_Nonclinical": "M2.6",
    "TM_5_Phase1": "M2.7",
    "TM_5_Phase2": "M2.7",
    "TM_5_Phase3": "M2.7",
}

def _normalize_section(s: str) -> str:
    s = (s or "").strip().upper()
    return s if s.startswith("M") else f"M{s}"

class RegulationRAGTool:
    """
    규제 검증·정규화·근거 반환.
    임계값 근거: CFG.COVERAGE_MIN, CFG.RAG_CONF_MIN, CFG.VIO_MAX, CFG.GENERATE_GATE
    """

    def __init__(
        self,
        auto_normalize: bool = True,
        max_violations: Optional[int] = None,
        coverage_threshold: Optional[float] = None,
        enable_rag: bool = True,
        llama_client=None,
    ):
        self.auto_normalize = auto_normalize
        self.max_violations = max_violations if max_violations is not None else CFG.VIO_MAX
        self.coverage_threshold = coverage_threshold if coverage_threshold is not None else CFG.COVERAGE_MIN
        self.enable_rag = enable_rag
        self.llama_client = llama_client

        self.mfds_rag: Optional[MFDSRAGTool] = None
        self.glossary_rag: Optional[GlossaryRAGTool] = None
        self.normalizer: Optional[TermNormalizer] = None

        if enable_rag:
            try:
                self.mfds_rag = MFDSRAGTool()
                self.glossary_rag = GlossaryRAGTool()
            except Exception:
                self.enable_rag = False

        if auto_normalize:
            try:
                self.normalizer = TermNormalizer(client=self.llama_client)
            except Exception:
                self.normalizer = None

    # -------- Excel 전체 검증 --------
    def validate_excel(self, excel_path: str, auto_fix: bool = True) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        results: List[Dict[str, Any]] = []
        total_violations = 0
        total_coverage = 0.0

        for sheet_name in wb.sheetnames:
            module = _normalize_section(SHEET_TO_MODULE.get(sheet_name, ""))
            if not module:
                continue

            ws = wb[sheet_name]
            content = self._extract_sheet_content(ws)
            if len(content) < 10:
                continue

            r = self.validate_and_normalize(section=module, content=content, auto_fix=auto_fix)
            r["sheet_name"] = sheet_name
            r["module"] = module
            results.append(r)

            total_violations += len(r["violations"])
            total_coverage += r["coverage"]

        validated_count = len(results)
        pass_count = sum(1 for r in results if r["pass"])
        return {
            "total_sheets": len(wb.sheetnames),
            "validated_sheets": validated_count,
            "results": results,
            "summary": {
                "total_violations": total_violations,
                "avg_coverage": (total_coverage / validated_count) if validated_count else 0.0,
                "pass_rate": (pass_count / validated_count) if validated_count else 0.0,
            },
        }

    def _extract_sheet_content(self, ws) -> str:
        lines = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append(" ".join(cells))
        return "\n".join(lines)

    # -------- 단일 섹션 검증 --------
    def validate_and_normalize(self, section: str, content: str, auto_fix: bool = True) -> Dict[str, Any]:
        section = _normalize_section(section)

        guideline_results: List[Dict[str, Any]] = []
        if self.enable_rag and self.mfds_rag:
            try:
                guideline_results = self.mfds_rag.search_by_module(query=content[:500], module=section, k=5)
            except Exception:
                guideline_results = []

        violations = self._detect_violations(content, guideline_results)

        normalized_content = content
        if auto_fix and violations and self.normalizer:
            try:
                normalized_content = self._normalize_content(content, violations)
            except Exception:
                normalized_content = content

        coverage = self._calculate_coverage(normalized_content, guideline_results)
        rag_conf = self._calculate_confidence(guideline_results)
        gloss = self._glossary_hit_rate(normalized_content)

        if coverage < self.coverage_threshold:
            guideline_results.extend(self._expand_coverage(section, normalized_content))
            coverage = self._calculate_coverage(normalized_content, guideline_results)

        citations = self._generate_citations(guideline_results)

        vio_w = self._violation_weight(violations)
        score_raw = 0.55 * coverage + 0.30 * rag_conf + 0.15 * gloss
        score = max(0.0, score_raw - 0.05 * vio_w)

        passed = (
            coverage >= CFG.COVERAGE_MIN and
            rag_conf >= CFG.RAG_CONF_MIN and
            vio_w <= CFG.VIO_MAX
        )

        return {
            "validated": True,
            "pass": passed,
            "violations": violations,
            "normalized_content": normalized_content,
            "coverage": coverage,
            "citations": citations,
            "rag_conf": rag_conf,
            "metrics": {
                "score": score,
                "score_raw": score_raw,
                "vio_weight": vio_w,
                "glossary_hit": gloss,
                "thresholds": {
                    "coverage_min": CFG.COVERAGE_MIN,
                    "rag_conf_min": CFG.RAG_CONF_MIN,
                    "vio_max": CFG.VIO_MAX,
                    "generate_gate": CFG.GENERATE_GATE,
                },
            },
        }

    # -------- Helpers --------
    def _detect_violations(self, content: str, guidelines: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        v: List[Dict[str, Any]] = []
        low = content.lower()
        placeholders = ["tbd", "to be defined", "to be decided", "미정", "lorem ipsum", "as appropriate", "etc."]
        if any(p in low for p in placeholders):
            v.append({
                "type": "placeholder",
                "description": "placeholder detected",
                "suggestion": "replace placeholders with actual values",
                "severity": "major",
            })
        return v

    def _normalize_content(self, content: str, violations: List[Dict[str, Any]]) -> str:
        if not self.normalizer:
            return content
        return self.normalizer.normalize(content)

    def _calculate_coverage(self, content: str, guidelines: List[Dict[str, Any]]) -> float:
        def toks(s: str) -> set[str]:
            return set(re.findall(r"[A-Za-z가-힣0-9]{3,}", s or ""))

        keys = toks(content)
        if not keys:
            return 0.0

        hits = 0
        for g in guidelines or []:
            txt = (g.get("content") or "") + " " + " ".join(map(str, (g.get("metadata") or {}).values()))
            hits += len(keys.intersection(toks(txt)))
        return min(1.0, hits / max(1, len(keys)))

    def _calculate_confidence(self, results: List[Dict[str, Any]]) -> float:
        if not results:
            return 0.0
        scores = [r.get("score", 0.0) for r in results[:3]]
        return sum(scores) / len(scores) if scores else 0.0

    def _glossary_hit_rate(self, content: str) -> float:
        if not self.glossary_rag:
            return 0.0
        try:
            hits = self.glossary_rag.search(content[:120]) or []
        except Exception:
            hits = []
        return 0.0 if not hits else min(1.0, sum(float(r.get("score", 0.0)) for r in hits) / len(hits))

    def _expand_coverage(self, section: str, content: str) -> List[Dict[str, Any]]:
        if not (self.enable_rag and self.mfds_rag):
            return []
        try:
            return self.mfds_rag.search_with_mmr(query=content[:500], k=3, fetch_k=10, lambda_mult=0.5)
        except Exception:
            return []

    def _generate_citations(self, guidelines: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        cites: List[Dict[str, Any]] = []
        for r in guidelines or []:
            md = r.get("metadata", {}) or {}
            cites.append({
                "source": md.get("source", "N/A"),
                "section": md.get("module", "N/A"),
                "page": md.get("page", "N/A"),
                "snippet": (r.get("content") or "")[:200] + "...",
                "score": r.get("score", 0.0)
            })
        return cites

    def _violation_weight(self, vlist: List[Dict[str, Any]]) -> int:
        w = {"minor": 1, "major": 2, "critical": 4}
        tot = 0
        for v in vlist:
            sev = str(v.get("severity", "major")).lower()
            tot += w.get(sev, 2)
        return tot

if __name__ == "__main__":
    tool = RegulationRAGTool(auto_normalize=False, enable_rag=False)
    sample = "임상은 다기관 무작위배정 이중맹검으로 수행되었다. TBD."
    out = tool.validate_and_normalize("M2.7", sample, auto_fix=False)
    print({"pass": out["pass"], "metrics": out["metrics"]})
