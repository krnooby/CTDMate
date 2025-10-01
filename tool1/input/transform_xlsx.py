import os, yaml, pandas as pd
from openpyxl import Workbook

INPUT_DIR = "../input"
OUTPUT_FILE = "./CTD_bundle_v2.xlsx"

wb = Workbook()
# ws = wb.active
# ws.title = "Index"
#ws.append(["Source File", "CTD Section", "Sheet Name"])

def detect_section(filename):
    import re
    m = re.search(r"(M2(?:\.\d+)*)", filename.replace("_","."))
    return m.group(1).upper() if m else "UNKNOWN"

for fname in os.listdir(INPUT_DIR):
    path = os.path.join(INPUT_DIR, fname)
    if not os.path.isfile(path):
        continue
    base, ext = os.path.splitext(fname)
   # section = detect_section(fname)
    sheet_name = base[:25]  # Excel 시트 이름 제한 (31자)
    #ws.append([fname, section, sheet_name])

    if ext == ".txt":
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.read().splitlines()
        ws = wb.create_sheet(sheet_name)
        ws.append(["Text"])  # ← 라인 번호 컬럼 제거, 단일 컬럼
        for line in lines:
            ws.append([line])

    elif ext.lower() in [".yaml", ".yml"]:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            data = yaml.safe_load(f)
        ws_new = wb.create_sheet(sheet_name)
        ws_new.append(["Text"])
        yaml_text = yaml.safe_dump(data, allow_unicode=True, sort_keys=False)
        for line in yaml_text.splitlines():
            ws_new.append([line])

wb.save(OUTPUT_FILE)
print(f"생성 완료 → {OUTPUT_FILE}")
