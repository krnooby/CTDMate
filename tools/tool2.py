"""
Tool2: Regulation RAG Tool

규제 체크리스트 검증 및 정규화를 수행하는 도구입니다.

주요 기능:
- ICH/MFDS 가이드라인 기반 검증
- 용어 자동 정규화 (major violations)
- Hybrid 검색 (Vector + BM25 + MMR)
- 규제 근거 citation 반환
- CTD_bundle.xlsx 파싱 및 시트별 검증

Architecture:
    Input (CTD_bundle.xlsx) → Parse sheets → Validate → Auto-normalize → Return citations

사용 예시:
    from tools.reg_rag import RegulationRAGTool

    # 방법 1: 엑셀 파일 전체 검증
    tool = RegulationRAGTool()
    results = tool.validate_excel("tool1/input/CTD_bundle.xlsx")

    # 방법 2: 특정 시트/섹션만 검증
    result = tool.validate_and_normalize(
        section="M2.7",
        content="임상시험 개요...",
        auto_fix=True
    )
"""

import sys
from pathlib import Path
from typing import Dict, List, Any, Optional
import openpyxl

# 프로젝트 루트 경로 추가
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from rag.mfds_rag import MFDSRAGTool
from rag.glossary_rag import GlossaryRAGTool
from rag.term_normalizer import TermNormalizer
# from rag.bm25_search import BM25Search  # TODO: 필요시 활성화


# CTD 시트명 → 모듈 매핑
SHEET_TO_MODULE = {
    "TM_5_M2_3_QOS": "M2.3",
    "TM_5_M2_4_Nonclinical_Ove": "M2.4",
    "TM_5_M2_5_Clinical_Overvi": "M2.5",
    "TM_5_M2_6_Nonclinical_Sum": "M2.6",
    "TM_5_M2_7_Clinical_Summar": "M2.7",
    "TM_5_Admin_Labeling_KR": "M1",
    "TM_5_Nonclinical": "M2.6",  # 비임상 원본 데이터
    "TM_5_Phase1": "M2.7",  # 임상 1상 데이터
    "TM_5_Phase2": "M2.7",  # 임상 2상 데이터
    "TM_5_Phase3": "M2.7",  # 임상 3상 데이터
}


class RegulationRAGTool:
    """
    Tool2: 규제 검증 및 정규화 도구

    Attributes:
        mfds_rag: MFDS/ICH 가이드라인 검색
        glossary_rag: 용어집 검색
        normalizer: 용어 정규화 (Llama-3.2)
        bm25: BM25 검색
    """

    def __init__(
        self,
        auto_normalize: bool = True,
        max_violations: int = 10,
        coverage_threshold: float = 0.7,
        enable_rag: bool = True
    ):
        """
        RegulationRAGTool 초기화

        Args:
            auto_normalize: 자동 정규화 활성화 여부
            max_violations: 최대 허용 위반 건수
            coverage_threshold: 최소 커버리지 비율 (0~1)
            enable_rag: RAG 도구 초기화 여부 (False면 경량 모드)
        """
        print("🔧 RegulationRAGTool 초기화 중...")

        self.auto_normalize = auto_normalize
        self.max_violations = max_violations
        self.coverage_threshold = coverage_threshold
        self.enable_rag = enable_rag

        # RAG 도구 초기화 (선택적)
        if enable_rag:
            try:
                self.mfds_rag = MFDSRAGTool()
                self.glossary_rag = GlossaryRAGTool()
                print("✅ RAG 도구 초기화 완료")
            except Exception as e:
                print(f"⚠️  RAG 도구 초기화 실패: {e}")
                print("   → 경량 모드로 동작합니다 (RAG 비활성화)")
                self.mfds_rag = None
                self.glossary_rag = None
                self.enable_rag = False
        else:
            self.mfds_rag = None
            self.glossary_rag = None

        # 용어 정규화기 (선택적)
        if auto_normalize:
            try:
                self.normalizer = TermNormalizer()
            except Exception as e:
                print(f"⚠️  TermNormalizer 초기화 실패: {e}")
                self.normalizer = None
        else:
            self.normalizer = None

        # BM25 검색기 (커버리지 확장용)
        self.bm25 = None  # TODO: BM25Searcher 초기화

        print("✅ RegulationRAGTool 초기화 완료\n")

    def validate_excel(
        self,
        excel_path: str,
        auto_fix: bool = True
    ) -> Dict[str, Any]:
        """
        CTD_bundle.xlsx 전체 검증

        Args:
            excel_path: 엑셀 파일 경로
            auto_fix: 자동 수정 활성화 여부

        Returns:
            {
                "total_sheets": int,
                "validated_sheets": int,
                "results": List[Dict],  # 시트별 검증 결과
                "summary": {
                    "total_violations": int,
                    "avg_coverage": float,
                    "pass_rate": float
                }
            }
        """
        print(f"📂 CTD Bundle 검증 시작: {excel_path}")
        print("=" * 70)

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        results = []
        total_violations = 0
        total_coverage = 0.0

        for sheet_name in wb.sheetnames:
            # CTD 모듈 매핑
            module = SHEET_TO_MODULE.get(sheet_name, "UNKNOWN")

            # 시트 내용 추출
            ws = wb[sheet_name]
            content = self._extract_sheet_content(ws)

            print(f"\n🔍 시트 검증: {sheet_name} → {module}")
            print(f"   내용 길이: {len(content)} 자")

            # 검증 수행
            if len(content) < 10:
                print(f"   ⚠️  내용이 너무 짧아 건너뜀")
                continue

            result = self.validate_and_normalize(
                section=module,
                content=content,
                auto_fix=auto_fix
            )

            # 결과 저장
            result['sheet_name'] = sheet_name
            result['module'] = module
            results.append(result)

            total_violations += len(result['violations'])
            total_coverage += result['coverage']

        # 요약 통계
        validated_count = len(results)
        pass_count = sum(1 for r in results if r['pass'])

        summary = {
            "total_sheets": len(wb.sheetnames),
            "validated_sheets": validated_count,
            "results": results,
            "summary": {
                "total_violations": total_violations,
                "avg_coverage": total_coverage / validated_count if validated_count > 0 else 0.0,
                "pass_rate": pass_count / validated_count if validated_count > 0 else 0.0
            }
        }

        print("\n" + "=" * 70)
        print("📊 검증 완료 요약")
        print("=" * 70)
        print(f"   - 검증 시트: {validated_count}/{len(wb.sheetnames)}")
        print(f"   - Pass 비율: {summary['summary']['pass_rate']:.1%}")
        print(f"   - 평균 커버리지: {summary['summary']['avg_coverage']:.1%}")
        print(f"   - 총 위반사항: {total_violations}개")
        print("=" * 70 + "\n")

        return summary

    def _extract_sheet_content(self, ws) -> str:
        """
        엑셀 시트의 텍스트 내용 추출

        Args:
            ws: openpyxl worksheet

        Returns:
            추출된 텍스트
        """
        lines = []
        for row in ws.iter_rows(values_only=True):
            # None이 아닌 셀만 추출
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if cells:
                lines.append(" ".join(cells))

        return "\n".join(lines)

    def validate_and_normalize(
        self,
        section: str,
        content: str,
        auto_fix: bool = True
    ) -> Dict[str, Any]:
        """
        규제 검증 및 자동 정규화

        Args:
            section: CTD 섹션 (예: "M2.7", "M2.6")
            content: 검증할 내용
            auto_fix: 자동 수정 활성화 여부

        Returns:
            {
                "validated": bool,
                "pass": bool,
                "violations": List[Dict],
                "normalized_content": str,
                "coverage": float,
                "citations": List[Dict],
                "rag_conf": float
            }
        """
        print(f"🔍 규제 검증 시작: {section}")

        # 1단계: 가이드라인 검색 (RAG 활성화 시)
        guideline_results = []
        if self.enable_rag and self.mfds_rag:
            try:
                guideline_results = self.mfds_rag.search_by_module(
                    query=content[:500],  # 첫 500자로 검색
                    module=section,
                    k=5
                )
            except Exception as e:
                print(f"⚠️  RAG 검색 실패: {e}")
        else:
            print(f"   → RAG 비활성화 모드 (가이드라인 검색 건너뜀)")

        # 2단계: 위반사항 탐지
        violations = self._detect_violations(content, guideline_results)

        # 3단계: 자동 정규화 (선택적)
        normalized_content = content
        if auto_fix and violations and self.normalizer:
            print(f"⚠️  {len(violations)}개 위반사항 발견, 자동 정규화 시도...")
            normalized_content = self._normalize_content(content, violations)

        # 4단계: 커버리지 계산
        coverage = self._calculate_coverage(normalized_content, guideline_results)

        # 5단계: 커버리지 확장 (BM25 + MMR)
        if coverage < self.coverage_threshold:
            print(f"📊 커버리지 {coverage:.2%} < {self.coverage_threshold:.2%}, 추가 검색...")
            expanded_results = self._expand_coverage(section, normalized_content)
            guideline_results.extend(expanded_results)
            coverage = self._calculate_coverage(normalized_content, guideline_results)

        # 6단계: Citation 생성
        citations = self._generate_citations(guideline_results)

        result = {
            "validated": True,
            "pass": len(violations) == 0,
            "violations": violations,
            "normalized_content": normalized_content,
            "coverage": coverage,
            "citations": citations,
            "rag_conf": self._calculate_confidence(guideline_results)
        }

        print(f"✅ 검증 완료: pass={result['pass']}, coverage={coverage:.2%}\n")
        return result

    def _detect_violations(
        self,
        content: str,
        guidelines: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        위반사항 탐지

        Args:
            content: 검증할 내용
            guidelines: 가이드라인 검색 결과

        Returns:
            위반사항 리스트 [{"type", "description", "suggestion"}]
        """
        violations = []

        # TODO: 실제 규칙 기반 검증 로직 구현
        # 예시:
        # - 필수 섹션 누락 체크
        # - 용어 표준화 체크
        # - 포맷 규칙 체크

        return violations

    def _normalize_content(
        self,
        content: str,
        violations: List[Dict[str, Any]]
    ) -> str:
        """
        용어/포맷 자동 정규화

        Args:
            content: 원본 내용
            violations: 위반사항 리스트

        Returns:
            정규화된 내용
        """
        if not self.normalizer:
            return content

        # TODO: TermNormalizer를 사용한 정규화
        normalized = self.normalizer.normalize(content)
        return normalized

    def _calculate_coverage(
        self,
        content: str,
        guidelines: List[Dict[str, Any]]
    ) -> float:
        """
        가이드라인 커버리지 계산

        Args:
            content: 검증 대상 내용
            guidelines: 가이드라인 검색 결과

        Returns:
            커버리지 점수 (0~1)
        """
        # TODO: 실제 커버리지 계산 로직
        # 예시: content의 주요 키워드가 guidelines에 얼마나 매칭되는지
        return 0.85

    def _expand_coverage(
        self,
        section: str,
        content: str
    ) -> List[Dict[str, Any]]:
        """
        BM25 + MMR로 커버리지 확장

        Args:
            section: CTD 섹션
            content: 검색 쿼리

        Returns:
            추가 검색 결과
        """
        # MMR 검색으로 다양성 확보
        mmr_results = self.mfds_rag.search_with_mmr(
            query=content[:500],
            k=3,
            fetch_k=10,
            lambda_mult=0.5
        )

        return mmr_results

    def _generate_citations(
        self,
        guidelines: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        Citation 메타데이터 생성

        Args:
            guidelines: 가이드라인 검색 결과

        Returns:
            Citation 리스트 [{"source", "section", "page", "snippet"}]
        """
        citations = []
        for result in guidelines:
            citations.append({
                "source": result['metadata'].get('source', 'N/A'),
                "section": result['metadata'].get('module', 'N/A'),
                "page": result['metadata'].get('page', 'N/A'),
                "snippet": result['content'][:200] + "...",
                "score": result.get('score', 0.0)
            })
        return citations

    def _calculate_confidence(
        self,
        results: List[Dict[str, Any]]
    ) -> float:
        """
        RAG 신뢰도 점수 계산

        Args:
            results: 검색 결과

        Returns:
            신뢰도 점수 (0~1)
        """
        if not results:
            return 0.0

        # 상위 결과들의 평균 유사도
        scores = [r.get('score', 0.0) for r in results[:3]]
        return sum(scores) / len(scores) if scores else 0.0


# 사용 예시 및 테스트
if __name__ == "__main__":
    print("=" * 70)
    print("  Tool2: Regulation RAG Tool 테스트")
    print("=" * 70)
    print()

    # 도구 초기화 (경량 모드: RAG 비활성화)
    tool = RegulationRAGTool(
        auto_normalize=False,
        enable_rag=False  # 테스트에서는 RAG 비활성화
    )

    # ========== 테스트 1: CTD_bundle.xlsx 전체 검증 ==========
    print("\n" + "=" * 70)
    print("  테스트 1: CTD_bundle.xlsx 전체 검증")
    print("=" * 70)

    excel_path = "../tool1/input/CTD_bundle.xlsx"
    if Path(excel_path).exists():
        summary = tool.validate_excel(excel_path, auto_fix=False)

        print("\n📊 시트별 검증 결과:")
        for i, result in enumerate(summary['results'][:5], 1):  # 처음 5개만 출력
            print(f"\n{i}. [{result['sheet_name']}] → {result['module']}")
            print(f"   - Pass: {result['pass']}")
            print(f"   - Coverage: {result['coverage']:.2%}")
            print(f"   - Violations: {len(result['violations'])}개")
            if result['citations']:
                print(f"   - Citations: {len(result['citations'])}개")
    else:
        print(f"⚠️  파일을 찾을 수 없습니다: {excel_path}")

    # ========== 테스트 2: 개별 시트 검증 ==========
    print("\n" + "=" * 70)
    print("  테스트 2: 개별 내용 검증")
    print("=" * 70)

    test_content = """
    meta:
      document: CTD M2.7 임상 요약 – TM-5(가상)
      language: ko
      version: '1.0'

    임상시험 개요

    본 임상시험은 Phase 2/3 다기관 무작위배정 이중맹검 위약대조 연구로 수행되었습니다.
    총 500명의 환자가 등록되었으며, 시험약군과 위약군에 1:1로 무작위 배정되었습니다.

    주요 효능 평가변수는 12주차의 증상 개선율이었으며, 시험약군에서 통계적으로 유의한
    개선을 보였습니다 (p<0.001).
    """

    result = tool.validate_and_normalize(
        section="M2.7",
        content=test_content,
        auto_fix=False
    )

    print("\n📋 검증 결과:")
    print(f"   - Pass: {result['pass']}")
    print(f"   - Coverage: {result['coverage']:.2%}")
    print(f"   - RAG Confidence: {result['rag_conf']:.2%}")
    print(f"   - Violations: {len(result['violations'])}개")
    print(f"   - Citations: {len(result['citations'])}개")

    print("\n" + "=" * 70)
    print("  테스트 완료")
    print("=" * 70)
