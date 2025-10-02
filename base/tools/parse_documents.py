from __future__ import annotations
from pathlib import Path
from typing import List, Dict, Any, Optional
import hashlib
import time
import json
import re
import openpyxl
import os

from langchain_core.tools import tool  # @tool(함수→도구, function calling용)

# ===== 고정 설정 (옵션 최소화) =====
OUTPUT_MD_DIR = Path("./out/md")
OUTPUT_RAG_DIR = Path("./out/rag")
SUPPORTED_EXTS = {".xlsx", ".pdf"}          # 확장자 유지
PDF_SPLIT_MODE = "page"                     # PDF 파싱 분할 단위
RAG_MAX_CHARS = 1600                        # RAG 청크 길이(문자)
RAG_OVERLAP = 200                           # RAG 청크 오버랩(문자)

# ========== 유틸 ==========
def _ensure_dirs(*dirs: Path):
    for d in dirs:
        d.mkdir(parents=True, exist_ok=True)

def _html_to_markdown(html: str) -> str:
    # HTML→Markdown 변환기(markdownify)
    from markdownify import markdownify as md
    return md(html or "")

def _sha256(s: str | bytes) -> str:
    h = hashlib.sha256()
    if isinstance(s, str):
        s = s.encode("utf-8")
    h.update(s)
    return h.hexdigest()

def _now_iso() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

def _approx_tokens(text: str) -> int:
    return max(1, int(len(text) / 4))

def _validate_input(p: Path):
    if not p.exists():
        raise FileNotFoundError(f"Input not found: {p}")
    if p.suffix.lower() not in SUPPORTED_EXTS:
        raise ValueError(f"Unsupported type: {p.name} (only .xlsx or .pdf)")

# ========== Upstage Document Parse ==========
def _parse_with_upstage(path: Path, split: str = "page"):
    # Upstage Document Parse(API, OCR(광학문자판독) 포함)
    from langchain_upstage import UpstageDocumentParseLoader
    api_key = os.getenv("UPSTAGE_API_KEY", "")
    loader = UpstageDocumentParseLoader(str(path), split=split, api_key=api_key)
    return loader.load()

# ========== Markdown → RAG 청크(헤딩 기반) ==========
HEADING_RX = re.compile(r"(?m)^(#{1,6})\s+(.*)$")

def _split_by_headings(md: str) -> list[dict]:
    spans = []
    matches = list(HEADING_RX.finditer(md))
    if not matches:
        return [{"level": 0, "title": "", "start": 0, "end": len(md)}]
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i+1].start() if i+1 < len(matches) else len(md)
        level = len(m.group(1))
        title = m.group(2).strip()
        spans.append({"level": level, "title": title, "start": start, "end": end})
    return spans

def _window_chunks(text: str, max_chars: int, overlap: int):
    if len(text) <= max_chars:
        yield (0, len(text)); return
    step = max(1, max_chars - overlap)
    i = 0
    while i < len(text):
        j = min(len(text), i + max_chars)
        yield (i, j)
        if j == len(text): break
        i += step

def chunk_markdown_for_rag(md: str, source_path: str, file_stem: str, page_hint: int | None = None) -> list[dict]:
    chunks = []
    headings = _split_by_headings(md)
    for span_idx, span in enumerate(headings):
        seg = md[span["start"]:span["end"]]
        for k, (a, b) in enumerate(_window_chunks(seg, max_chars=RAG_MAX_CHARS, overlap=RAG_OVERLAP)):
            text = seg[a:b].strip()
            if not text:
                continue
            chunk_id = f"{file_stem}::p{page_hint or 0}::h{span_idx}::w{k}"
            chunks.append({
                "id": _sha256(source_path + chunk_id + text)[:16],
                "text": text,
                "metadata": {
                    "source": source_path,
                    "file_name": file_stem,
                    "page": page_hint,
                    "heading_level": span["level"],
                    "heading": span["title"],
                    "chunk_index": k,
                    "span_index": span_idx,
                    "created_at": _now_iso(),
                    "char_len": len(text),
                    "approx_tokens": _approx_tokens(text),
                }
            })
    return chunks

# ========== XLSX → 시트 1청크(JSON 객체) ==========
def _ws_to_markdown(ws) -> str:
    rows = []
    for r in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c not in (None, "") else " " for c in r]
        rows.append(cells)
    if not rows:
        return ""
    maxc = max(len(r) for r in rows)
    rows = [r + [""] * (maxc - len(r)) for r in rows]
    md = []
    md.append(f"# Sheet: {ws.title}")
    header = rows[0]
    md.append("| " + " | ".join(header) + " |")
    md.append("| " + " | ".join(["---"] * len(header)) + " |")
    for r in rows[1:]:
        md.append("| " + " | ".join(r) + " |")
    return "\n".join(md)

def xlsx_to_sheet_chunks(source_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(source_path, data_only=True)
    items = []
    for idx, name in enumerate(wb.sheetnames):
        ws = wb[name]
        text = _ws_to_markdown(ws).strip()
        if not text:
            continue
        chunk_id = f"{Path(source_path).stem}::sheet::{idx}"
        items.append({
            "id": _sha256(source_path + chunk_id + text)[:16],
            "text": text,
            "metadata": {
                "source": source_path,
                "file_name": Path(source_path).stem,
                "sheet_name": name,
                "sheet_index": idx,
                "created_at": _now_iso(),
                "char_len": len(text),
                "approx_tokens": _approx_tokens(text),
            }
        })
    return items

# ========== 파서 파이프라인(run) ==========
def run(inputs: List[str]) -> dict:
    _ensure_dirs(OUTPUT_MD_DIR, OUTPUT_RAG_DIR)

    results: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    for raw in inputs:
        try:
            src = Path(raw)
            _validate_input(src)

            md_pages: List[str] = []
            rag_items: List[dict] = []

            if src.suffix.lower() == ".xlsx":
                # (A) XLSX: 시트 단위 JSONL, MD는 Upstage로 파싱(원형 보존)
                rag_items = xlsx_to_sheet_chunks(str(src))
                docs = _parse_with_upstage(src, split=PDF_SPLIT_MODE)
                for d in docs:
                    md_pages.append(_html_to_markdown(d.page_content or ""))
            else:
                # (B) PDF: 페이지→MD→헤딩 기반 청크
                docs = _parse_with_upstage(src, split=PDF_SPLIT_MODE)
                for i, d in enumerate(docs, 1):
                    md_page = _html_to_markdown(d.page_content or "")
                    md_pages.append(md_page)
                    rag_items.extend(
                        chunk_markdown_for_rag(
                            md=md_page,
                            source_path=str(src),
                            file_stem=src.stem,
                            page_hint=i
                        )
                    )

            # MD 저장
            merged_md = "\n\n---\n\n".join(md_pages) if md_pages else ""
            md_path = OUTPUT_MD_DIR / f"{src.stem}.md"
            md_path.write_text(merged_md, encoding="utf-8")

            # JSONL 저장
            rag_path = OUTPUT_RAG_DIR / f"{src.stem}.jsonl"
            with rag_path.open("w", encoding="utf-8") as f:
                for item in rag_items:
                    text = " ".join(item["text"].splitlines())
                    text = re.sub(r"\|\s+\|", "| |", text)
                    text = re.sub(r"\|\s*\|", "", text)
                    item["text"] = text.strip()
                    f.write(json.dumps(item, ensure_ascii=False) + "\n")

            results.append({
                "input": str(src),
                "markdown": str(md_path),
                "rag_jsonl": str(rag_path),
                "pages": len(md_pages),
            })

        except Exception as e:
            errors.append({"input": raw, "error": str(e)})

    return {"ok": len(errors) == 0, "results": results, "errors": errors}

# ========== Tool adapter(@tool) ==========
@tool("parse_documents")
def parse_documents(file_paths: Optional[List[str]] = None,
                    texts: Optional[List[str]] = None) -> str:
    """
    파서 툴 어댑터(adapter, 어댑터). run() 호출 결과를 메인과 호환되는 JSON 문자열로 반환.
    반환 스키마(schema, 스키마):
      {"documents":[{"source","markdown",...},...], "combined_markdown": str, "extracted": {}}
    """
    items: List[Dict] = []

    # 파일 입력
    if file_paths:
        res = run(file_paths)
        for r in res.get("results", []):
            src = r.get("input", "")
            md_path = Path(r.get("markdown", ""))
            md_text = md_path.read_text(encoding="utf-8") if md_path.is_file() else ""
            items.append({"source": src, "markdown": md_text, "meta": {"pages": r.get("pages", 0)}})

    # 인라인 텍스트 입력
    if texts:
        for i, t in enumerate(texts, 1):
            items.append({"source": f"inline_{i}", "markdown": t, "meta": {"mode": "inline"}})

    combined = "\n\n---\n\n".join([it["markdown"] for it in items]) if items else ""
    payload = {
        "documents": items,
        "combined_markdown": combined,
        "extracted": {}  # 필요 시 Information Extract(정보 추출) 결과 주입 가능
    }
    return json.dumps(payload, ensure_ascii=False)
